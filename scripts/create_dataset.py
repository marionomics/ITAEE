"""
create_dataset.py

Builds a unified quarterly dataset combining:
  - ITAEE (Indicador Trimestral de Actividad Económica Estatal) from INEGI
  - Exchange rates vs MXN from Yahoo Finance

Output: data/dataset.csv
"""

import os
import warnings
import openpyxl
import pandas as pd
import yfinance as yf

warnings.filterwarnings("ignore", category=FutureWarning)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ITAEE_STATES_FILE = os.path.join(BASE_DIR, "data", "tabulados_ITAEE", "ITAEE_3.xlsx")
ITAEE_NATIONAL_FILE = os.path.join(BASE_DIR, "data", "tabulados_ITAEE", "ITAEE_20.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "data", "dataset.csv")

TARGET_STATES = [
    "Aguascalientes",
    "Durango",
    "Querétaro",
    "San Luis Potosí",
    "Zacatecas",
]

# Direct MXN pairs available on Yahoo Finance
DIRECT_MXN_PAIRS = {
    "USD": "USDMXN=X",
    "CAD": "CADMXN=X",
    "BRL": "BRLMXN=X",
    "JPY": "JPYMXN=X",
}

# Currencies that need cross-rate via USD (no direct MXN pair on Yahoo)
# We download XXXUSD=X (or USDXXX=X) and multiply by USDMXN
CROSS_VIA_USD = {
    "CNY": "USDCNY=X",   # invert: USDMXN / USDCNY
    "COP": "USDCOP=X",   # invert: USDMXN / USDCOP
    "CLP": "USDCLP=X",   # invert: USDMXN / USDCLP
    "RUB": "USDRUB=X",   # invert: USDMXN / USDRUB
    "ILS": "USDILS=X",   # invert: USDMXN / USDILS
    "EUR": "EURUSD=X",   # multiply: EURUSD * USDMXN
    "INR": "USDINR=X",   # invert: USDMXN / USDINR
}


# ---------------------------------------------------------------------------
# 1. Parse ITAEE
# ---------------------------------------------------------------------------
def _parse_quarter_cols(ws):
    """Extract (col_index, year, quarter) tuples from the header rows."""
    row5 = [c.value for c in list(ws.iter_rows(min_row=5, max_row=5))[0]]
    row6 = [c.value for c in list(ws.iter_rows(min_row=6, max_row=6))[0]]

    current_year = None
    quarter_cols = []
    for i, (yr, q) in enumerate(zip(row5, row6)):
        if yr is not None and yr != "Concepto":
            current_year = str(yr).replace("R", "").replace("P", "")
        if q in ("T1", "T2", "T3", "T4") and current_year:
            quarter_cols.append((i, current_year, q))
    return quarter_cols


def _make_period_index(quarter_cols):
    """Build a PeriodIndex from quarter_cols."""
    quarter_map = {"T1": 1, "T2": 2, "T3": 3, "T4": 4}
    return [
        pd.Period(year=int(yr), quarter=quarter_map[q], freq="Q")
        for _, yr, q in quarter_cols
    ]


def load_itaee() -> pd.DataFrame:
    """Read ITAEE state data and national total into a single DataFrame."""
    print("Loading ITAEE data...")

    # --- State-level data from ITAEE_3.xlsx ---
    wb = openpyxl.load_workbook(ITAEE_STATES_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    quarter_cols = _parse_quarter_cols(ws)

    state_data = {}
    for row in ws.iter_rows(min_row=8, max_row=40, values_only=False):
        name = row[0].value
        if name and name.strip() in TARGET_STATES:
            state_data[name.strip()] = [
                row[col_idx].value for col_idx, _, _ in quarter_cols
            ]
    wb.close()

    # --- National-level data from ITAEE_20.xlsx (row 8 = "Total") ---
    wb_nat = openpyxl.load_workbook(ITAEE_NATIONAL_FILE, read_only=True)
    ws_nat = wb_nat[wb_nat.sheetnames[0]]
    qc_nat = _parse_quarter_cols(ws_nat)

    for row in ws_nat.iter_rows(min_row=8, max_row=8, values_only=False):
        state_data["Nacional"] = [
            row[col_idx].value for col_idx, _, _ in qc_nat
        ]
    wb_nat.close()

    # --- Build DataFrame ---
    periods = _make_period_index(quarter_cols)
    df = pd.DataFrame(state_data, index=periods)
    df.index.name = "quarter"
    df = df.apply(pd.to_numeric, errors="coerce")
    df = df.dropna(how="all")

    # Reorder: Nacional first, then states alphabetically
    state_cols = sorted([c for c in df.columns if c != "Nacional"])
    df = df[["Nacional"] + state_cols]

    df.columns = ["ITAEE_" + col for col in df.columns]

    print(f"  ITAEE range: {df.index[0]} to {df.index[-1]} ({len(df)} quarters)")
    print(f"  Series: {list(df.columns)}")
    return df


# ---------------------------------------------------------------------------
# 2. Download exchange rates from Yahoo Finance
# ---------------------------------------------------------------------------
def load_exchange_rates(start: str, end: str) -> pd.DataFrame:
    """Download daily exchange rates and compute cross-rates vs MXN.

    Strategy:
      - Direct pairs (USD, CAD, BRL, JPY): download XXXMXN=X
      - Cross-rate pairs: download USDMXN=X + USDXXX=X, then compute
        XXXMXN = USDMXN / USDXXX  (or XXXUSD * USDMXN for EUR)
    All series are resampled to quarterly averages.
    """
    print("Downloading exchange rates from Yahoo Finance...")

    # Collect all tickers we need
    all_tickers = list(DIRECT_MXN_PAIRS.values()) + list(CROSS_VIA_USD.values())
    # Ensure USDMXN is included (needed for cross-rates and is also a target)
    if "USDMXN=X" not in all_tickers:
        all_tickers.append("USDMXN=X")
    all_tickers = list(set(all_tickers))

    data = yf.download(all_tickers, start=start, end=end, progress=False)

    if isinstance(data.columns, pd.MultiIndex):
        close = data["Close"]
    else:
        close = data

    usdmxn = close["USDMXN=X"]

    # Build result DataFrame with XXXMXN rates
    fx = pd.DataFrame(index=close.index)

    # Direct pairs
    for label, ticker in DIRECT_MXN_PAIRS.items():
        fx[label] = close[ticker]

    # Cross-rate pairs
    for label, ticker in CROSS_VIA_USD.items():
        if label == "EUR":
            # EURUSD * USDMXN = EURMXN
            fx[label] = close[ticker] * usdmxn
        else:
            # USDMXN / USDXXX = XXXMXN
            fx[label] = usdmxn / close[ticker]

    # Resample to quarterly average
    df_q = fx.resample("QE").mean()

    # Convert DatetimeIndex to PeriodIndex
    df_q.index = df_q.index.to_period("Q")
    df_q.index.name = "quarter"

    # Prefix columns
    df_q.columns = ["FX_" + col for col in df_q.columns]

    print(f"  FX range: {df_q.index[0]} to {df_q.index[-1]} ({len(df_q)} quarters)")
    return df_q


# ---------------------------------------------------------------------------
# 3. Merge and save
# ---------------------------------------------------------------------------
def main():
    df_itaee = load_itaee()

    # Download FX for the ITAEE date range
    start_date = df_itaee.index[0].start_time.strftime("%Y-%m-%d")
    end_date = df_itaee.index[-1].end_time.strftime("%Y-%m-%d")
    df_fx = load_exchange_rates(start_date, end_date)

    # Inner join on quarterly period — keeps only overlapping quarters
    df = df_itaee.join(df_fx, how="inner")

    # Drop rows with any NaN to get a clean balanced panel for VAR
    df_full = df.dropna()

    print(f"\nRaw merged: {df.shape[0]} quarters x {df.shape[1]} variables")
    print(f"After dropping NaNs: {df_full.shape[0]} quarters")
    print(f"Period: {df_full.index[0]} to {df_full.index[-1]}")
    print(f"\nColumns: {list(df_full.columns)}")

    df = df_full

    # Save
    df.to_csv(OUTPUT_FILE)
    print(f"\nDataset saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
